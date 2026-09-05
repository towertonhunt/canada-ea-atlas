#!/usr/bin/env python3
"""Promote discovered proponent-site documents onto the map and into the archive.

Reads the discovery inventories (data/raw/proponents/sites/*.json from the
live crawl, data/raw/proponents/wayback/*.json from the Internet Archive
index), keeps what looks like EA record material, dedupes against every
document the map already links, and:

  * attaches a document to an existing map project of the same proponent
    when that project's name appears in the document's URL or title
    (-> data/raw/proponent_docs_index.json, merged by build_national_geojson)
  * otherwise files it in a per-proponent "EA document library" feature
    (-> data/raw/proponent_libraries.json + data/docs/proponent/library-<key>.json)

Original URLs stay primary; a Wayback capture is carried as `fallback_url`
so archive_docs.py can mirror the file even when the proponent has removed
it. --report writes an .xlsx of every discovered document with the decision
taken, for review.

  python3 scripts/promote_proponent_docs.py --report /path/review.xlsx
"""
import argparse
import collections
import glob
import json
import os
import re
import sys
import urllib.parse

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, 'data', 'raw')
PROP = os.path.join(RAW, 'proponents')
OUT_INDEX = os.path.join(RAW, 'proponent_docs_index.json')
OUT_LIBS = os.path.join(RAW, 'proponent_libraries.json')
OUT_DOCS = os.path.join(ROOT, 'data', 'docs', 'proponent')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from build_proponent_targets import norm as norm_name  # noqa: E402

EA_TYPES = {'ESR_final', 'ESR_draft', 'ESR', 'EA_report', 'terms_of_reference',
            'notice_of_commencement', 'notice_of_completion', 'screening', 'open_house',
            'monitoring', 'closure', 'technical', 'permit_approval', 'engagement'}
# 'other' documents are still promoted when the URL/title carries a clear EA signal
EA_SIGNAL = re.compile(r'environment|assessment|\besr\b|\beis\b|class[-_ ]?ea|impact[-_ ]?statement|'
                       r'follow[-_ ]?up|reclamation|decommission|mitigation|baseline|'
                       r'terms[-_ ]?of[-_ ]?reference|screening|open[-_ ]?house|consultation', re.I)
MEDIA = re.compile(r'\.(jpe?g|png|gif|svg|webp|mp4|mp3|css|js|ico|woff2?)(\?|$)', re.I)
# Proponents whose corporate site is global: keep only Canada-related material
GLOBAL_HOSTS = {'bhp.com', 'equinor.com', 'sasol.com', 'glencore.com', 'riotinto.com',
                'newmont.com', 'shell.com', 'conocophillips.com', 'imperialoil.com'}
CANADA = re.compile(r'canad|ontario|alberta|british[-_ ]?columbia|\bbc\b|quebec|qu[eé]bec|'
                    r'manitoba|saskatchewan|labrador|newfoundland|nova[-_ ]?scotia|'
                    r'new[-_ ]?brunswick|nunavut|yukon|nwt|northwest[-_ ]?territor|jansen|'
                    r'potash|ekati|athabasca|oil[-_ ]?sands|kearl|cold[-_ ]?lake|sarnia|'
                    r'kitimat|lng[-_ ]?canada|bay[-_ ]?du[-_ ]?nord|hebron|hibernia|terra[-_ ]?nova',
                    re.I)
STOP = {'project', 'projects', 'transmission', 'line', 'station', 'generating', 'generation',
        'expansion', 'upgrade', 'replacement', 'phase', 'power', 'energy', 'plant', 'facility',
        'mine', 'mining', 'reinforcement', 'redevelopment', 'refurbishment', 'development',
        'environmental', 'assessment', 'study', 'report', 'ontario', 'canada', 'north', 'south',
        'east', 'west', 'lake', 'river', 'creek', 'area', 'system', 'works', 'water', 'hydro',
        'electric', 'nuclear', 'wind', 'solar', 'gas', 'pipeline', 'terminal', 'road', 'highway',
        'bridge', 'unit', 'units', 'application', 'modernization', 'international', 'impacting',
        'construction', 'installation', 'improvements', 'improvement', 'rehabilitation',
        'maintenance', 'operations', 'operation', 'management', 'monitoring', 'proposed',
        'program', 'programme', 'services', 'service', 'building', 'property', 'centre',
        'center', 'district', 'regional', 'municipal', 'county', 'township', 'community',
        'northern', 'southern', 'eastern', 'western', 'central', 'district', 'storage',
        'supply', 'treatment', 'wastewater', 'sewage', 'landfill', 'quarry', 'gravel'}


def norm_url(u):
    p = urllib.parse.urlsplit(u.strip())
    host = (p.hostname or '').lower()
    host = host[4:] if host.startswith('www.') else host
    path = urllib.parse.unquote(p.path).rstrip('/').lower()
    q = p.query.lower()
    return f'{host}{path}' + (f'?{q}' if q else '')


def tokens(name):
    return {t for t in re.findall(r'[a-z0-9]+', name.lower()) if len(t) >= 5 and t not in STOP}


def load_existing_urls():
    seen = set()
    for path in glob.glob(os.path.join(ROOT, 'data', 'docs', '*', '*.json')):
        if '/proponent/' in path:
            continue
        try:
            for d in json.load(open(path)).get('docs', []):
                if d.get('url'):
                    seen.add(norm_url(d['url']))
        except (ValueError, OSError):
            pass
    return seen


def load_inventories():
    """-> {key: {'name':..., 'docs': [rec...]}} merging live + wayback per proponent."""
    inv = {}
    for path in glob.glob(os.path.join(PROP, 'sites', '*.json')):
        d = json.load(open(path))
        e = inv.setdefault(d['key'], {'name': d.get('name') or d['key'], 'host': d.get('host'),
                                      'docs': []})
        for x in d['docs']:
            e['docs'].append({'url': x['url'], 'title': x.get('title') or '', 'type': x.get('type'),
                              'source': 'live', 'wayback_url': None, 'timestamp': None,
                              'bytes': None, 'noise': False, 'page': x.get('page')})
    for path in glob.glob(os.path.join(PROP, 'wayback', '*.json')):
        d = json.load(open(path))
        e = inv.setdefault(d['key'], {'name': d.get('name') or d['key'], 'host': d.get('host'),
                                      'docs': []})
        for x in d['docs']:
            e['docs'].append({'url': x['url'], 'title': x.get('title') or '', 'type': x.get('type'),
                              'source': 'wayback', 'wayback_url': x.get('wayback_url'),
                              'timestamp': x.get('timestamp'), 'bytes': x.get('bytes'),
                              'noise': bool(x.get('noise')), 'page': None})
    return inv


def project_index(geo, targets):
    """-> {proponent key: [(feature_key, name, token set, docs_path)]}"""
    by_key = collections.defaultdict(list)
    variants = {}
    for t in targets:
        for v in t.get('variants', []) + [t['name']]:
            variants[norm_name(v)] = t['key']
    for f in geo['features']:
        p = f['properties']
        prop = p.get('proponent')
        if not prop:
            continue
        k = variants.get(norm_name(prop))
        if not k:
            for vk, key in variants.items():
                if len(vk) >= 6 and vk in norm_name(prop):
                    k = key
                    break
        if not k:
            continue
        toks = tokens(p.get('name') or '')
        if len(toks) >= 1:
            by_key[k].append((f"{p.get('source')}|{p.get('name')}", p.get('name'), toks,
                              p.get('docs_path')))
    # a token that appears in many of one proponent's project names (e.g.
    # "transformer") cannot identify a project on its own
    freq = {}
    for k, lst in by_key.items():
        c = collections.Counter(t for _, _, toks, _ in lst for t in toks)
        freq[k] = c
    return by_key, freq


def decide(rec, host, existing, seen_run):
    """-> (promote: bool, reason)"""
    blob = f"{rec['title']} {urllib.parse.unquote(rec['url'])}"
    if MEDIA.search(rec['url']):
        return False, 'media file'
    if rec['noise']:
        return False, 'noise (annual/investor/newsletter/fr)'
    apex = host[4:] if host and host.startswith('www.') else host
    if apex in GLOBAL_HOSTS and not CANADA.search(blob):
        return False, 'global site, no Canada signal'
    if rec['type'] not in EA_TYPES and not EA_SIGNAL.search(blob):
        return False, 'not EA-typed and no EA signal'
    nu = norm_url(rec['url'])
    if nu in existing:
        return False, 'already on the map (registry catalogue)'
    if nu in seen_run:
        return False, 'duplicate within discovery'
    return True, 'promoted'


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--report', help='write an .xlsx review workbook here')
    args = ap.parse_args()

    geo = json.load(open(os.path.join(ROOT, 'data', 'projects_canada.geojson')))
    targets = json.load(open(os.path.join(PROP, 'targets.json')))
    tmeta = {t['key']: t for t in targets}
    existing = load_existing_urls()
    inv = load_inventories()
    projects, tokfreq = project_index(geo, targets)

    index = collections.defaultdict(list)     # feature_key -> docs
    libraries = {}                            # proponent key -> {name, docs}
    rows = []
    stats = collections.Counter()
    for key, e in sorted(inv.items()):
        seen_run = set()
        # prefer the live record when the same URL appears in both sources
        e['docs'].sort(key=lambda r: (r['source'] != 'live', r['url']))
        for rec in e['docs']:
            ok, reason = decide(rec, e.get('host') or '', existing, seen_run)
            matched = None
            if ok:
                seen_run.add(norm_url(rec['url']))
                # match on what names the document -- its title and file name --
                # not on directory names in the path ("/lower-mainland/" would
                # otherwise tie every regional file to one project)
                upath = urllib.parse.unquote(urllib.parse.urlsplit(rec['url']).path.rstrip('/'))
                fname = os.path.basename(upath)
                blob = f"{rec['title']} {fname}".lower().replace('_', ' ').replace('-', ' ')
                # exact folder names still count for a single, highly distinctive
                # token: Hydro One files live in /majorprojects/Waasigan/Documents/
                # with generic titles. "/lower-mainland/" does not equal "mainland".
                segs = {seg.lower().replace('_', ' ') for seg in upath.split('/')[:-1]}
                best, best_n = None, 0
                fq = tokfreq.get(key, {})
                for fk, name, toks, dp in projects.get(key, []):
                    hit = [t for t in toks if re.search(rf'\b{re.escape(t)}\b', blob)]
                    distinctive = [t for t in hit if fq.get(t, 0) <= 2]
                    n = len(hit)
                    ok_match = (n >= 2 and distinctive) or \
                               (n == 1 and len(hit[0]) >= 8 and fq.get(hit[0], 0) == 1)
                    if not ok_match:
                        folder = [t for t in toks if t in segs and len(t) >= 8 and fq.get(t, 0) == 1]
                        if folder:
                            hit, n, ok_match = folder, 1, True
                    if ok_match and n > best_n:
                        best, best_n = (fk, name, dp), n
                entry = {'title': rec['title'] or os.path.basename(rec['url']),
                         'category': rec['type'] or 'other', 'url': rec['url'],
                         'source': 'proponent_site' if rec['source'] == 'live' else 'wayback'}
                if rec['wayback_url']:
                    entry['fallback_url'] = rec['wayback_url']
                if rec['timestamp']:
                    entry['captured'] = rec['timestamp'][:8]
                if best:
                    matched = best[1]
                    index[best[0]].append(entry)
                    stats['matched to project'] += 1
                else:
                    lib = libraries.setdefault(key, {'name': e['name'], 'docs': []})
                    lib['docs'].append(entry)
                    stats['library'] += 1
            stats[reason] += 1
            rows.append([e['name'], e.get('host'), rec['source'], rec['type'], rec['title'][:250],
                         rec['url'], rec['wayback_url'], rec['timestamp'][:8] if rec['timestamp'] else None,
                         rec['bytes'], 'yes' if ok else 'no', reason, matched])

    os.makedirs(OUT_DOCS, exist_ok=True)
    for old in glob.glob(os.path.join(OUT_DOCS, 'library-*.json')):
        os.remove(old)
    libs_out = []
    for key, lib in libraries.items():
        fn = f'library-{re.sub(r"[^a-z0-9]+", "-", key)}.json'
        json.dump({'project': f"{lib['name']} EA document library", 'docs': lib['docs']},
                  open(os.path.join(OUT_DOCS, fn), 'w'), ensure_ascii=False)
        t = tmeta.get(key, {})
        libs_out.append({'key': key, 'name': f"{lib['name']}: EA document library",
                         'proponent': lib['name'], 'website': t.get('website'),
                         'doc_count': len(lib['docs']), 'docs_path': f'data/docs/proponent/{fn}'})
    json.dump(libs_out, open(OUT_LIBS, 'w'), ensure_ascii=False, indent=1)
    json.dump({k: v for k, v in index.items()}, open(OUT_INDEX, 'w'), ensure_ascii=False)
    print(f'{len(rows)} discovered documents; decisions: {dict(stats.most_common())}')
    print(f'-> {len(index)} projects gain documents; {len(libs_out)} proponent libraries')

    if args.report:
        from openpyxl import Workbook
        wb = Workbook(write_only=True)
        ws = wb.create_sheet('documents')
        ws.append(['proponent', 'host', 'found_via', 'type', 'title', 'url', 'wayback_url',
                   'captured', 'bytes', 'promoted', 'reason', 'matched_project'])
        for r in rows:
            ws.append(r)
        ws2 = wb.create_sheet('summary')
        ws2.append(['proponent', 'discovered', 'promoted', 'matched_to_project', 'library'])
        per = collections.defaultdict(collections.Counter)
        for r in rows:
            per[r[0]]['discovered'] += 1
            if r[9] == 'yes':
                per[r[0]]['promoted'] += 1
                per[r[0]]['matched' if r[11] else 'library'] += 1
        for name, c in sorted(per.items(), key=lambda x: -x[1]['promoted']):
            ws2.append([name, c['discovered'], c['promoted'], c['matched'], c['library']])
        ws3 = wb.create_sheet('decision_key')
        for k, v in stats.most_common():
            ws3.append([k, v])
        wb.save(args.report)
        print('report ->', args.report)


if __name__ == '__main__':
    main()
